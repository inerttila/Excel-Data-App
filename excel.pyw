import datetime
import os
import tkinter as tk
import tkinter.messagebox as messagebox
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from tkcalendar import Calendar
from backup_excel import create_backup
from buttons import create_buttons
from copyserver import copy_to_server
from weekly_totals import calculate_weekly_total_hours
from qrcode_module import generate_qr_code_and_start_server
# from email_sender import send_email
from tkinter import (Entry, Label, OptionMenu, StringVar, Text, Tk,
                     Toplevel, simpledialog, ttk)


# Define file paths as constants
base_dir = os.path.dirname(os.path.abspath(__file__))
excel_file_name = "Timesheet-managementt.xlsx"
excel_file_path = os.path.join(base_dir, excel_file_name)

# Constants for local and server file paths
local_file_path = excel_file_path
server_file_path = r"\\192.168.40.21\\Fileshare SV1\\Timesheet-managementt.xlsx"


# Calculate the weekly total hours by summing the hours in the worksheet
def display_weekly_total():
    global worksheet
    weekly_total_hours = calculate_weekly_total_hours(worksheet)

    # Show the total hours message as a messagebox
    weekly_total_message = f"Total Hours: {weekly_total_hours} hours"
    messagebox.showinfo("Total Hours", weekly_total_message)


# Function to apply header styles to a worksheet
def apply_header_styles(worksheet, headers):
    header_font = Font(bold=True, color="ffffff")
    header_fill = PatternFill(
        start_color="061c43", end_color="061c43", fill_type="solid"
    )

    for col_num, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        # Apply right alignment to numeric columns (e.g., 'Hours')
        row_height = 30
        worksheet.row_dimensions[1].height = row_height


# Function to create or get a sheet from a workbook
def create_or_get_sheet(workbook, sheet_name):
    try:
        return workbook[sheet_name]
    except KeyError:
        return workbook.create_sheet(sheet_name)


# Function to create a directory if it doesn't exist
def create_directory_if_not_exists(file_path):
    directory = os.path.dirname(file_path)
    os.makedirs(directory, exist_ok=True)


# Create a new workbook if it doesn't exist
def load_or_create_workbook(file_path):
    try:
        return openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        # Save the workbook
        workbook.save(file_path)

        return workbook


# Function to handle user input for categories with options
def select_option(category, options):
    try:
        value = simpledialog.askstring("Data Input",
                                       f"Enter {category}:")
        if value in options:
            input_values[category].set(value)
        else:
            raise ValueError("Invalid option selected.")
    except ValueError as e:
        messagebox.showerror("Error", str(e))


# Function to handle selection of a date
def select_date():
    if hasattr(select_date, "calendar_open") and select_date.calendar_open:
        return

    select_date.calendar_open = True

    def on_date_selected():
        selected_date = cal.selection_get()
        input_values["Date"].set(selected_date.strftime("%Y-%m-%d"))
        date_label.config(text=input_values["Date"].get())
        top.destroy()
        select_date.calendar_open = False

    # Create a pop-up window for date selection
    top = Toplevel()
    top.title("Select Date")
    top.configure(background="#000000")

    # Set the icon for the pop-up window
    icon = tk.PhotoImage(
        file="C:\\Users\\User\\Desktop\\Excel-Data-App\\Media\\otr.png")
    top.iconphoto(False, icon)

    style = ttk.Style(top)
    style.configure(
        "Calendar.Treeview",
        background="#000000",
        foreground="#FCFCFC",
        fieldbackground="#000000"
    )

    style.configure(
        "TButton", foreground="black", background="#061c43",
        font=("Arial", 9, "bold")
    )
    # Create a calendar widget for date selection
    calendar_position_row = 1
    calendar_position_column = 1
    cal = Calendar(top)

    cal.grid(
        row=calendar_position_row, column=calendar_position_column,
        padx=10, pady=10
    )

    # Create a "Confirm" button for date selection
    confirm_button = ttk.Button(
        top, text="Confirm", command=on_date_selected, style="Custom.TButton"
    )
    confirm_button.grid(
        row=calendar_position_row + 1, column=calendar_position_column, padx=10, pady=10
    )

    # Function to close the pop-up window

    def close_window():
        top.destroy()
        select_date.calendar_open = False

    # Configure the behavior of the close button
    top.protocol("WM_DELETE_WINDOW", close_window)

    # Start the main event loop for the pop-up window
    top.mainloop()


# Function to validate hours input
def validate_hours_input(value):
    if value and not value.replace(".", "").isdigit():
        return False
    return True


# Function to validate notes input
def validate_notes_input(value):
    return True


# Function to check if the last entry in the worksheet is on a Fiday
def is_last_entry_on_sunday(worksheet):
    if worksheet.max_row <= 1:
        return False  # No data in the worksheet

    last_date_value = worksheet.cell(row=worksheet.max_row, column=1).value
    if last_date_value is not None:
        last_date_value = datetime.datetime.strptime(
            last_date_value, "%Y-%m-%d")
        # Friday is represented as 4 in the weekday() function
        return last_date_value.weekday() == 4
    return False
    


def confirm_input():
    global workbook
    try:
        row_data = []
        for category in categories:
            value = input_values[category].get()
            if category == "Date" and not value:
                raise ValueError("Date must be selected.")
            if category == "Hours":
                if not value:
                    messagebox.showerror("Error", "Please write the hours.")
                    return
                if not validate_hours_input(value):
                    messagebox.showerror("Error", "Invalid Hours input.")
                    return
            if category == "Notes":
                value = entry.get("1.0", "end-1c")
            row_data.append(value)

        # Convert the hours input to a float
        row_data[5] = float(row_data[5])

        # Calculate the total hours per week
        weekly_total_hours = calculate_weekly_total_hours(worksheet)

        # Check if the weekly total hours will exceed 40 after adding the new entry
        if weekly_total_hours + int(row_data[5]) > 40:
            if not is_last_entry_on_sunday(worksheet):
                raise ValueError("Total weekly hours cannot exceed 40 hours.")
            else:
                # Prompt the user for confirmation
                user_choice = messagebox.askyesno(
                    "Weekly Hours Exceeded",
                    "Adding this entry will clear the the weekly 'Sheet' . Do you want to continue?",
                )
                if user_choice:
                    # User chose to continue, add an empty row before appending new data on Monday
                    sheet_name_to_delete = "Sheet"
                    delete_sheet("s", sheet_name_to_delete)
                    header_row = [
                        "Date",
                        "Service Line",
                        "Type of Service",
                        "Company",
                        "Task",
                        "Hours",
                        "Notes",
                    ]
                    create_sheet_with_headers(header_row, row_data)
                else:
                    # User chose not to continue, clear the input fields
                    for category in categories:
                        input_values[category].set("")
                    return  # Exit the function without saving

        # Create or get the "2024" sheet and save data there too
        sheet_2024 = create_or_get_sheet(workbook, "2024")
        set_column_widths(sheet_2024)
        today = datetime.date.today()
        is_monday = today.weekday() == 0
        # Check if the last entry in the original sheet is on Sunday
        if is_last_entry_on_sunday(worksheet) and is_monday:
            # Add an empty row before appending new data on Monday
            sheet_name_to_delete = "Sheet"
            delete_sheet("s", sheet_name_to_delete)
            header_row = [
                "Date",
                "Service Line",
                "Type of Service",
                "Company",
                "Task",
                "Hours",
                "Notes",
            ]
            create_sheet_with_headers(header_row, row_data)
            sheet_2024.append([])

        worksheet.append(row_data)
        sheet_2024.append(row_data)
        workbook.save(excel_file_path)

        # Reset input values and user interface elements
        for category in categories:
            input_values[category].set("")

            if category == "Notes":
                entry.delete(1.0, "end")

            if category == "Company":
                reset_option_menu(company_option_menu, company_options)
            elif category == "Service Line":
                reset_option_menu(service_line_option_menu, service_options)
            elif category == "Type of Service":
                reset_option_menu(type_of_service_option_menu,
                                  type_of_service_options)
            elif category == "Task":
                reset_option_menu(task_option_menu, task_options)

        messagebox.showinfo("Success", "Data saved successfully.")

    except Exception as e:
        messagebox.showerror("Error", str(e))


# Function to reset an option menu with new options
def reset_option_menu(option_menu, options):
    menu = option_menu["menu"]
    menu.delete(0, "end")
    for option in options:
        menu.add_command(
            label=option,
            command=lambda value=option: input_values[option_menu.category].set(
                value),
        )
    input_values[option_menu.category].set(options[0])


# Check if the file exists
if os.path.isfile(excel_file_path):
    # The file exists, load it
    workbook = load_workbook(excel_file_path)
else:
    # The file doesn't exist, create it
    workbook = openpyxl.Workbook()
    workbook.save(excel_file_path)

# File path and initial data setup
workbook = load_workbook(local_file_path)
categories = [
    "Date",
    "Service Line",
    "Type of Service",
    "Company",
    "Task",
    "Hours",
    "Notes",
]
company_options = [
    "SKAITECH",
    "3DSKAI",
    "-"
]
service_options = [
    "Develop",
    "Drones/Robotics",
    "IoT/Automation",
    "Security/AI",
    "Immersive Technologies",
    "Training",
    "Research",
    "-",
]
type_of_service_options = [
    "Software",
    "Hardware",
    "Other",
    "-"
]
task_options = [
    "Development",
    "Control",
    "Research",
    "Testing",
    "Other",
    "-"
]
input_values = {}


# Function to set column widths for a sheet
def set_column_widths(sheet):
    column_widths = {
        "A": 15,  # Date
        "B": 20,  # Service Line
        "C": 20,  # Type of Service
        "D": 20,  # Company
        "E": 20,  # Task
        "F": 10,  # Hours
        "G": 100,  # Notes
    }

    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width


# Create necessary directories, load or create workbook and worksheet
create_directory_if_not_exists(excel_file_path)
workbook = load_or_create_workbook(excel_file_path)
worksheet = workbook["Sheet"]

# Set column widths
set_column_widths(workbook.active)

# Apply header styles to the worksheet
apply_header_styles(worksheet, categories)


# Open the Excel file using the default program
def open_excel_file():
    try:
        os.startfile(excel_file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Error opening Excel file: {str(e)}")


# Function to delete a sheet from the workbook
def delete_sheet(wb, sheet_name):
    try:
        global workbook
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            workbook.remove(sheet)
            workbook.save(excel_file_path)
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


# Function to create a new sheet with headers and data, and adjust column widths
def create_sheet_with_headers(headers, rd):
    try:
        global workbook
        new_sheet = workbook.create_sheet(title="Sheet", index=0)
        apply_header_styles(new_sheet, headers)

        # Set the column widths for the new sheet
        column_widths = {
            "A": 15,  # Date
            "B": 20,  # Service Line
            "C": 20,  # Type of Service
            "D": 20,  # Company
            "E": 20,  # Task
            "F": 10,  # Hours
            "G": 100,  # Notes
        }
        for column, width in column_widths.items():
            new_sheet.column_dimensions[column].width = width

        # workbook.save(excel_file_path)
        new_sheet.append(rd)

        # Save the workbook after appending data
        workbook.save(excel_file_path)
    except Exception:
        pass


# Create the main application window
window = Tk()
window.title("Timesheet Application")
window.configure(background="#000000")
icon = tk.PhotoImage(
    file="C:\\Users\\User\\Desktop\\Excel-Data-App\\Media\\otr.png")
window.iconphoto(False, icon)

# Set the dimensions and position of the window
window_width = 780
window_height = 300
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
x_coordinate = (screen_width - window_width) // 2
y_coordinate = (screen_height - window_height) // 2
window.geometry(
    f"{window_width}x{window_height}+{x_coordinate}+{y_coordinate}")

# Create labels and input fields for each category
for i, category in enumerate(categories):
    label = Label(window, text=category,
                  background="#000000", foreground="#FCFCFC")
    pady_value = 6
    label.grid(row=i, column=0, padx=(10, 20), pady=(pady_value, pady_value))

    # Create input fields based on the category
    if category == "Company":
        input_values[category] = StringVar(window, company_options[0])
        company_option_menu = OptionMenu(
            window, input_values[category], *company_options
        )
        company_option_menu.config(width=26)
        company_option_menu.category = category
        company_option_menu.grid(row=i, column=1)
    elif category == "Service Line":
        input_values[category] = StringVar(window, service_options[0])
        service_line_option_menu = OptionMenu(
            window, input_values[category], *service_options
        )
        service_line_option_menu.config(width=26)
        service_line_option_menu.category = category
        service_line_option_menu.grid(row=i, column=1)
    elif category == "Type of Service":
        input_values[category] = StringVar(window, type_of_service_options[0])
        type_of_service_option_menu = OptionMenu(
            window, input_values[category], *type_of_service_options
        )
        type_of_service_option_menu.config(width=26)
        type_of_service_option_menu.category = category
        type_of_service_option_menu.grid(row=i, column=1)
    elif category == "Task":
        input_values[category] = StringVar(window, task_options[0])
        task_option_menu = OptionMenu(
            window, input_values[category], *task_options)
        task_option_menu.config(width=26)
        task_option_menu.category = category
        task_option_menu.grid(row=i, column=1)
    elif category == "Date":
        input_values[category] = StringVar(window)
        date_label = Label(
            window, textvariable=input_values[category], bg="#000000", fg="#FCFCFC")
        date_label.grid(row=i, column=1)

    elif category == "Hours":
        input_values[category] = StringVar(window)
        vcmd = (window.register(validate_hours_input), "%P")
        entry = Entry(
            window,
            textvariable=input_values[category],
            validate="key",
            validatecommand=vcmd,
            font=("Arial", 10),
            width=28,
        )
        entry.grid(row=i, column=1, sticky="w")
    elif category == "Notes":
        input_values[category] = StringVar(window)
        vcmd = (window.register(validate_notes_input), "%P")
        entry = Text(
            window,
            height=4,
            width=28,
            font=("Arial", 10),
        )
        entry.grid(row=i, column=1, columnspan=2, sticky="w")
    else:
        input_values[category] = StringVar(window)
        entry = Entry(window, textvariable=input_values[category])
        entry.grid(row=i, column=1, sticky="w")

# Calculate the maximum width among button labels
button_labels = ["Confirm", "Send File",
                 "Open File", "Total", "Select Date", "Backup", "send_email" "Generate QR Code"]
max_button_width = max(len(label) for label in button_labels)

# Call the create_buttons function to create the buttons
buttons = create_buttons(
    window, categories, max_button_width, select_date, confirm_input, copy_to_server, open_excel_file, display_weekly_total,
    create_backup, generate_qr_code_and_start_server)  # Add "send_email" here

# Start the main event loop for the application window
window.mainloop()
